# excel_parser.py - LPS Book Excel Parser
import openpyxl
from typing import List, Dict, Any
from app.utils.helpers import log_info, log_error


def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    try:
        log_info(f"Parsing Excel: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Try LPS_DATA sheet first, then OVERALL, then first sheet
        if "LPS_DATA" in wb.sheetnames:
            sheet = wb["LPS_DATA"]
        elif "OVERALL" in wb.sheetnames:
            sheet = wb["OVERALL"]
        else:
            sheet = wb.worksheets[0]

        log_info(f"Reading sheet: {sheet.title}")

        # Read headers from row 1
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        log_info(f"Headers found: {headers}")

        # Normalize header → column index map
        col_map = {_norm(h): i for i, h in enumerate(headers)}

        # Find columns flexibly
        col_sf       = _find(col_map, "sf_no","sfno","sf","surveyno","survey")
        col_owner    = _find(col_map, "owner_name","owner","name","pattadar")
        col_district = _find(col_map, "district")
        col_taluk    = _find(col_map, "taluk")
        col_village  = _find(col_map, "village")
        col_cor_sqm  = _find(col_map, "corridor_sqm","corridorsqm","corridorsqm","corridor")
        col_cor_cen  = _find(col_map, "corridor_cent","corridorcent")
        col_tow_sqm  = _find(col_map, "tower_area_sqm","towerareasqm","towerarea")
        col_tow_cen  = _find(col_map, "tower_area_cent","towerareacent")
        col_tower    = _find(col_map, "tower_no","towerno","tower")
        col_loc_frm  = _find(col_map, "loc_from","locfrom")
        col_loc_to   = _find(col_map, "loc_to","locto")
        col_bn       = _find(col_map, "boundary_n","boundaryn")
        col_bs       = _find(col_map, "boundary_s","boundarys")
        col_be       = _find(col_map, "boundary_e","boundarye")
        col_bw       = _find(col_map, "boundary_w","boundaryw")

        log_info(f"sf:{col_sf} owner:{col_owner} corridor:{col_cor_sqm} tower:{col_tower}")

        parcels = []
        for row in list(sheet.iter_rows(values_only=True))[1:]:  # skip header
            if not row or all(v is None for v in row):
                continue

            sf_val = _str(row, col_sf)
            if not sf_val or sf_val.lower() in ("sf_no","none","nan",""):
                continue

            parcels.append({
                "sf_no":           sf_val,
                "owner":           _str(row, col_owner)    or "Unknown",
                "district":        _str(row, col_district) or "",
                "taluk":           _str(row, col_taluk)    or "",
                "village":         _str(row, col_village)  or "",
                "corridor_sqm":    _flt(row, col_cor_sqm),
                "corridor_cent":   _flt(row, col_cor_cen),
                "tower_area_sqm":  _flt(row, col_tow_sqm),
                "tower_area_cent": _flt(row, col_tow_cen),
                "tower_no":        _str(row, col_tower)    or "-",
                "loc_from":        _str(row, col_loc_frm)  or "",
                "loc_to":          _str(row, col_loc_to)   or "",
                "boundaries": {
                    "N": _str(row, col_bn) or "",
                    "S": _str(row, col_bs) or "",
                    "E": _str(row, col_be) or "",
                    "W": _str(row, col_bw) or "",
                }
            })

        log_info(f"Parsed {len(parcels)} parcels")
        return parcels

    except Exception as e:
        log_error("Excel parsing failed", e)
        return []


def _norm(s: str) -> str:
    return str(s).lower().replace(" ","").replace("_","").replace("-","")

def _find(col_map: dict, *candidates):
    for c in candidates:
        if _norm(c) in col_map:
            return col_map[_norm(c)]
    return None

def _str(row, idx) -> str:
    if idx is None or idx >= len(row): return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""

def _flt(row, idx) -> float:
    if idx is None or idx >= len(row): return 0.0
    try: return float(str(row[idx]).replace(",",""))
    except: return 0.0